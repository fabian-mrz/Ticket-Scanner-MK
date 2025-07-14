import sys
import os
import time  # Add this import
import sqlite3
import logging
import tempfile
import uvicorn
from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, RedirectResponse
from pydantic import BaseModel
from typing import List, Optional
from import_send import send_email_with_attachments, create_ticket, import_tickets
import random
from openpyxl import load_workbook
import pandas as pd
from datetime import datetime
from openpyxl.styles import Border, Side, PatternFill
import io
from fastapi import FastAPI, HTTPException, UploadFile, File, Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import pandas as pd
from datetime import datetime, timedelta
from fastapi import HTTPException, Depends
from fastapi.security import OAuth2PasswordBearer
from jose import JWTError, jwt
from pydantic import BaseModel
from fastapi.security import OAuth2PasswordRequestForm
from fastapi.responses import JSONResponse
from fastapi import Depends
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.responses import HTMLResponse
from fastapi import Header
from pathlib import Path


# Security configuration
SECRET_KEY = "pleasechangeme" 
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30

# Get the project root directory
META_DB = "../data/meta.db"
TICKETS_DB = "../data/tickets.db"

USERS = {
    "admin": {
        "username": "admin",
        "password": "admin123"  
    }
}

# Add these new functions for authentication
def create_access_token(username: str):
    expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    return jwt.encode(
        {"sub": username, "exp": expire}, 
        SECRET_KEY, 
        algorithm=ALGORITHM
    )


oauth2_scheme = OAuth2PasswordBearer(tokenUrl="login")

def verify_token(authorization: str = Header(None)):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(
            status_code=401,
            detail="No valid token provided"
        )
    try:
        token = authorization.split(" ")[1]
        jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        return True
    except JWTError:
        raise HTTPException(
            status_code=401,
            detail="Invalid token"
        )

# Setup logging file ticket_system.log
log_file = Path("ticket_system.log")
if not log_file.exists():
    log_file.touch()

# Configure logging
logging.basicConfig(
    filename=log_file,
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)



app = FastAPI()

app.mount("/static2", StaticFiles(directory="static2"), name="static2")

class Ticket(BaseModel):
    name_ticketinhaber: str
    email_ticketinhaber: str
    email_kaeufer: str
    # Make other fields optional since we don't need them for updates
    ticket_id: Optional[str] = None
    ticket_number: Optional[str] = None
    bestell_id: Optional[str] = None
    ticket_type: Optional[str] = None
    discounted: Optional[bool] = None

class NumberRange(BaseModel):
    pool_name: str
    range_start: Optional[int] = None
    range_end: Optional[int] = None
    specific_numbers: Optional[List[int]] = None

# Add new Pydantic model for manual ticket creation
class ManualTicket(BaseModel):
    ticket_type: str  # "Tribüne", "Sitz-/ Stehplatz", etc.
    discounted: bool
    name_ticketinhaber: str
    email_ticketinhaber: str
    email_kaeufer: str
    bestell_id: str

class TicketLimit(BaseModel):
    key: str
    value: int

@app.post("/api/login")
async def login(credentials: dict):
    user = USERS.get(credentials.get("username"))
    if not user or credentials.get("password") != user["password"]:
        raise HTTPException(
            status_code=401,
            detail="Incorrect username or password"
        )
    return {
        "access_token": create_access_token(user["username"]),
        "token_type": "bearer"
    }

# Limits
@app.get("/config/limits")
async def get_ticket_limits(token: str = Depends(verify_token)):
    try:
        with sqlite3.connect(META_DB) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT key, value FROM config_values")
            limits = cursor.fetchall()
            return {
                "limits": [
                    {"key": key, "value": value} 
                    for key, value in limits
                ]
            }
    except Exception as e:
        raise HTTPException(
            status_code=500, 
            detail=f"Error fetching limits: {str(e)}"
        )

@app.put("/config/limits/{key}")
async def update_ticket_limit(
    key: str, 
    limit: TicketLimit, 
    token: str = Depends(verify_token)
):
    if key != limit.key:
        raise HTTPException(
            status_code=400, 
            detail="Path key and body key must match"
        )
        
    valid_keys = [
        "trib_normal_max", 
        "trib_discount_max", 
        "sitz_normal_max", 
        "sitz_discount_max"
    ]
    
    if key not in valid_keys:
        raise HTTPException(
            status_code=400, 
            detail=f"Invalid key. Must be one of: {', '.join(valid_keys)}"
        )

    try:
        with sqlite3.connect(META_DB) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE config_values 
                SET value = ? 
                WHERE key = ?
            """, (limit.value, key))
            
            if cursor.rowcount == 0:
                cursor.execute("""
                    INSERT INTO config_values (key, value)
                    VALUES (?, ?)
                """, (key, limit.value))
                
            conn.commit()
            return {"status": "success", "message": f"Updated {key} to {limit.value}"}
            
    except Exception as e:
        raise HTTPException(
            status_code=500, 
            detail=f"Error updating limit: {str(e)}"
        )
    
#Number Pools

@app.delete("/pools/{pool_name}/{number}/remove")
async def remove_number(pool_name: str, number: int, token: str = Depends(verify_token)):
    try:
        with sqlite3.connect(META_DB) as conn:
            cursor = conn.cursor()
            
            # Check if number is used
            cursor.execute("""
                SELECT used FROM number_pools 
                WHERE pool_name = ? AND number = ?
            """, (pool_name, number))
            
            result = cursor.fetchone()
            if not result:
                raise HTTPException(status_code=404, detail="Number not found")
            
            if result[0]:  # If number is used
                raise HTTPException(
                    status_code=400, 
                    detail="Cannot remove used number"
                )
            
            # Remove the number
            cursor.execute("""
                DELETE FROM number_pools 
                WHERE pool_name = ? AND number = ?
            """, (pool_name, number))
            
            conn.commit()
            return {"status": "success"}
            
    except sqlite3.Error as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/")
async def root():
    return RedirectResponse(url="/static2/login.html")

@app.get("/pools/{pool_name}")
async def get_pool_numbers(pool_name: str):
    try:
        with sqlite3.connect(META_DB) as conn, sqlite3.connect(TICKETS_DB) as tconn:
            cursor = conn.cursor()
            tcursor = tconn.cursor()
            cursor.execute("""
                SELECT number, used 
                FROM number_pools 
                WHERE pool_name = ?
                ORDER BY number
            """, (pool_name,))
            results = cursor.fetchall()
            numbers = []
            # Determine table name for tickets
            ticket_table = "trib" if pool_name == "TRIB_START" else "sitz"
            for r in results:
                number, used = r
                tcursor.execute(f"SELECT Einchecken FROM {ticket_table} WHERE TicketNumber = ?", (number,))
                ticket_row = tcursor.fetchone()
                einchecken = ticket_row[0] if ticket_row and ticket_row[0] else ""
                numbers.append({"number": number, "used": used, "einchecken": einchecken})
            return {"numbers": numbers}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/pools/update")
async def update_pool(range_data: NumberRange, token: str = Depends(verify_token)):
    try:
        with sqlite3.connect(META_DB) as conn:
            cursor = conn.cursor()
            
            # Clear existing unused numbers
            cursor.execute("""
                DELETE FROM number_pools 
                WHERE pool_name = ? AND used = FALSE
            """, (range_data.pool_name,))
            
            # Add range numbers
            if range_data.range_start and range_data.range_end:
                for num in range(range_data.range_start, range_data.range_end + 1):
                    cursor.execute("""
                        INSERT OR IGNORE INTO number_pools (pool_name, number)
                        VALUES (?, ?)
                    """, (range_data.pool_name, num))
            
            # Add specific numbers
            if range_data.specific_numbers:
                for num in range_data.specific_numbers:
                    cursor.execute("""
                        INSERT OR IGNORE INTO number_pools (pool_name, number)
                        VALUES (?, ?)
                    """, (range_data.pool_name, num))
            
            conn.commit()
            return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    

    
@app.get("/tickets")
async def get_tickets(token: str = Depends(verify_token)):
    try:
        with sqlite3.connect(TICKETS_DB) as conn:
            cursor = conn.cursor()
            tickets = []
            for table in ['trib', 'sitz']:
                cursor.execute(f"""
                    SELECT 
                        TicketNumber, qrcode, Ticket, Einchecken,
                        Bestell_ID, Bestellstatus, Ticket_ID,
                        Name_Ticketinhaber, Email_Ticketinhaber,
                        Name_Kaeufer, Email_Kaeufer, Discounted,
                        created_at
                    FROM {table}
                    ORDER BY created_at DESC
                """)
                tickets.extend([{
                    'table': table,
                    'ticket_number': row[0],
                    'qrcode': row[1],
                    'ticket_type': row[2],
                    'einchecken': row[3],
                    'bestell_id': row[4],
                    'bestellstatus': row[5],
                    'ticket_id': row[6],
                    'name_ticketinhaber': row[7],
                    'email_ticketinhaber': row[8],
                    'name_kaeufer': row[9],
                    'email_kaeufer': row[10],
                    'discounted': bool(row[11]),
                    'created_at': row[12]
                } for row in cursor.fetchall()])
            return {"tickets": tickets}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/tickets/resend/{ticket_id}")
async def resend_ticket(ticket_id: str, token: str = Depends(verify_token)):
    try:
        with sqlite3.connect(TICKETS_DB) as conn:
            cursor = conn.cursor()
            
            # Find ticket in either table
            for table in ['trib', 'sitz']:
                cursor.execute(f"""
                    SELECT * FROM {table}
                    WHERE Ticket_ID = ?
                """, (ticket_id,))
                ticket = cursor.fetchone()
                if ticket:
                    break
            
            if not ticket:
                raise HTTPException(status_code=404, detail="Ticket not found")
            
            # Create temporary folder for ticket
            with tempfile.TemporaryDirectory() as tmp_dir:
                ticket_path = os.path.join(tmp_dir, f"{ticket[0]}.jpg")
                
                # Recreate ticket
                create_ticket(
                    ticket[1],  # qrcode
                    ticket[2],  # ticket_type
                    ticket[0],  # ticket_number
                    ticket_path
                )
                
                # Send email
                send_email_with_attachments(
                    ticket[10],  # email_kaeufer
                    ticket[4],   # bestell_id
                    tmp_dir
                )
                
            return {"status": "success"}
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    


@app.put("/tickets/{ticket_id}")
async def update_ticket(ticket_id: str, ticket: Ticket, token: str = Depends(verify_token)):
    try:
        with sqlite3.connect(TICKETS_DB) as conn:
            cursor = conn.cursor()
            
            updated = False
            for table in ['trib', 'sitz']:
                try:
                    cursor.execute(f"""
                        UPDATE {table}
                        SET Name_Ticketinhaber = ?,
                            Email_Ticketinhaber = ?,
                            Email_Kaeufer = ?
                        WHERE Ticket_ID = ?
                    """, (
                        ticket.name_ticketinhaber,
                        ticket.email_ticketinhaber,
                        ticket.email_kaeufer,
                        ticket_id
                    ))
                    if cursor.rowcount > 0:
                        updated = True
                        conn.commit()
                        break
                except Exception as e:
                    logging.error(f"Error updating {table}: {e}")
                    continue
            
            if not updated:
                raise HTTPException(
                    status_code=404, 
                    detail=f"Ticket {ticket_id} not found"
                )
                
            return {"status": "success"}
            
    except Exception as e:
        logging.error(f"Error updating ticket {ticket_id}: {e}")
        raise HTTPException(
            status_code=500, 
            detail=f"Error updating ticket: {str(e)}"
        )
    
@app.post("/tickets/upload")
async def upload_tickets(file: UploadFile = File(...)):
    try:
        # Create temporary file to store the uploaded CSV
        with tempfile.NamedTemporaryFile(delete=False, suffix='.csv') as temp_file:
            content = await file.read()
            temp_file.write(content)
            temp_path = temp_file.name

        try:
            # Import tickets using the existing function
            tickets_created = import_tickets(temp_path, TICKETS_DB)
            os.unlink(temp_path)  # Clean up temp file
            
            return {
                "status": "success",
                "message": f"Successfully imported {tickets_created} tickets"
            }
        except Exception as e:
            logging.error(f"Error importing tickets: {e}")
            if os.path.exists(temp_path):
                os.unlink(temp_path)
            raise HTTPException(
                status_code=500,
                detail=f"Error importing tickets: {str(e)}"
            )
            
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error processing upload: {str(e)}"
        )

@app.post("/tickets/create")
async def create_manual_ticket(ticket: ManualTicket, token: str = Depends(verify_token)):
    try:
        with sqlite3.connect(TICKETS_DB) as conn:
            cursor = conn.cursor()
            
            # Determine table based on ticket type
            table_name = "trib" if "Tribüne" in ticket.ticket_type else "sitz"
            
            # Get next available ticket number
            cursor.execute(f"SELECT MAX(TicketNumber) FROM {table_name}")
            last_number = cursor.fetchone()[0] or 0
            ticket_number = int(last_number) + 1
            
            # Format ticket number based on type and if discounted
            formatted_ticket = f"{'E' if ticket.discounted else ''}{'TRI0' if table_name == 'trib' else 'STEH0'}{ticket_number}"
            
            # Generate QR code with random prefix
            random_code = str(random.randint(1000000, 9999999))
            qrcode_value = f"{random_code}_{formatted_ticket}"
            
            # Create temporary directory for ticket
            with tempfile.TemporaryDirectory() as tmp_dir:
                ticket_path = os.path.join(tmp_dir, f"{formatted_ticket}.jpg")
                
                # Create ticket image
                if create_ticket(qrcode_value, ticket.ticket_type, str(ticket_number), ticket_path):
                    # Insert into database with correct format
                    cursor.execute(f"""
                        INSERT INTO {table_name} (
                            TicketNumber, qrcode, Ticket, Einchecken, 
                            Bestell_ID, Bestellstatus, Ticket_ID,
                            Name_Ticketinhaber, Email_Ticketinhaber,
                            Name_Kaeufer, Email_Kaeufer, Discounted
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        ticket_number,
                        qrcode_value,
                        ticket.ticket_type,
                        "",  # Einchecken
                        ticket.bestell_id,
                        "Abgeschlossen",  # Bestellstatus
                        str(int(time.time())),  # Generate unique Ticket_ID
                        ticket.name_ticketinhaber,
                        ticket.email_ticketinhaber,
                        ticket.name_ticketinhaber,  # Same as ticketinhaber
                        ticket.email_kaeufer,
                        1 if ticket.discounted else 0  # Store as integer
                    ))
                    
                    # Send email
                    send_email_with_attachments(
                        ticket.email_kaeufer,
                        ticket.bestell_id,
                        tmp_dir
                    )
                    
                    conn.commit()
                    return {
                        "status": "success",
                        "ticket_number": formatted_ticket,
                        "message": "Ticket created and sent successfully"
                    }
                else:
                    raise HTTPException(
                        status_code=500,
                        detail="Failed to create ticket image"
                    )
                    
    except Exception as e:
        logging.error(f"Error creating manual ticket: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Error creating ticket: {str(e)}"
        )
    
def get_ticket_display_type(ticket_type: str) -> str:
    if "Tribüne" in ticket_type:
        return "Tribüne ermäßigt" if "ermäßigt" in ticket_type else "Tribüne"
    else:
        return "Sitz-/ Stehplatz ermäßigt" if "ermäßigt" in ticket_type else "Sitz-/ Stehplatz"
    
def get_ticket_price(ticket_type: str) -> str:
    if "Tribüne" in ticket_type:
        return "24,00 €" if "ermäßigt" in ticket_type else "20,00 €"
    else:  # Sitz-/Stehplatz
        return "18,00 €" if "ermäßigt" in ticket_type else "12,00 €"

@app.post("/tickets/convert-to-vvk")
async def convert_to_vvk(token: str = Depends(verify_token)):
    try:
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        
        # Set up header
        ws['A1'] = 'VVK    Waldburg-Tattoo 2025'
        ws['F2'] = 'Summe:'
        ws['G2'] = '- €'
        ws['A3'] = 'Anzahl:'
        ws['A4'] = '950'
        
        # Set column headers
        headers = [
            'Karten-#', 'Tribüne/Standard', 'Wo', 'Vorname', 'Name', 'Preis', 
            'Wann', 'Rech-#', 'email', 'Telefon', 'Art', 'Auftrag bestätigt',
            'Zahlung eingegangen', 'erledigt', 'Adresse', 'PLZ', 'Ort', 'Bemerkung'
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=5, column=col, value=header)
        
        # Get data from database
        with sqlite3.connect(TICKETS_DB) as conn:
            cursor = conn.cursor()
            row_num = 6  # Start after headers
            
            # Process tickets
            cursor.execute("""
                SELECT TicketNumber, Ticket, Name_Ticketinhaber, Email_Ticketinhaber, 
                    Bestell_ID, Discounted, created_at
                FROM trib
                ORDER BY TicketNumber
            """)
            trib_tickets = cursor.fetchall()

            cursor.execute("""
                SELECT TicketNumber, Ticket, Name_Ticketinhaber, Email_Ticketinhaber, 
                    Bestell_ID, Discounted, created_at
                FROM sitz
                ORDER BY TicketNumber
            """)
            sitz_tickets = cursor.fetchall()
            
            # Combine and process all tickets
            for ticket in trib_tickets + sitz_tickets:
                created_at = datetime.fromisoformat(ticket[6]) if ticket[6] else None
                created_at_str = created_at.strftime('%d.%m.%Y') if created_at else ''
                
                ticket_num = ticket[0]
                ticket_type = ticket[1]
                name_parts = ticket[2].split(' ', 1) if ticket[2] else ['', '']
                vorname = name_parts[0]
                nachname = name_parts[1] if len(name_parts) > 1 else ''
                
                # Fill row data
                ws.cell(row=row_num, column=1, value=f"{'E' if ticket[5] else ''}{'TRI' if 'Tribüne' in ticket_type else 'STEH'}{str(ticket_num).zfill(4)}")
                ws.cell(row=row_num, column=2, value=get_ticket_display_type(ticket_type))
                ws.cell(row=row_num, column=3, value='Online')
                ws.cell(row=row_num, column=4, value=vorname)
                ws.cell(row=row_num, column=5, value=nachname)
                ws.cell(row=row_num, column=6, value=get_ticket_price(ticket_type))  # Add price
                ws.cell(row=row_num, column=8, value=ticket[4])  # Bestell_ID
                ws.cell(row=row_num, column=9, value=ticket[3])  # Email
                ws.cell(row=row_num, column=7, value=created_at_str)  # Add created_at in 'Wann' column
                
                row_num += 1
        
        # Save to temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            wb.save(temp_file.name)
            temp_path = temp_file.name
            
        # Read file and prepare response
        with open(temp_path, 'rb') as f:
            contents = f.read()
        os.unlink(temp_path)  # Clean up
        
        return Response(
            content=contents,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                'Content-Disposition': 'attachment; filename="VVK_2025.xlsx"'
            }
        )
            
    except Exception as e:
        logging.error(f"Error converting to VVK: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Error converting to VVK: {str(e)}"
        )
 

# Add this endpoint
@app.get("/api/logs")
async def get_logs(token: str = Depends(verify_token)):
    try:
        log_file = Path("ticket_system.log")
        if not log_file.exists():
            return {"logs": []}
        
        with open(log_file, "r") as f:
            # Read last 1000 lines (adjust as needed)
            lines = f.readlines()[-1000:]
            
        # Parse log lines into structured format
        parsed_logs = []
        for line in lines:
            try:
                # Parse log line: "2024-04-26 10:15:30,123 - INFO - Message"
                parts = line.split(" - ", 2)
                if len(parts) == 3:
                    timestamp, level, message = parts
                    parsed_logs.append({
                        "timestamp": timestamp.strip(),
                        "level": level.strip(),
                        "message": message.strip()
                    })
            except Exception:
                continue
                
        return {"logs": parsed_logs}
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error reading logs: {str(e)}"
        )

@app.get("/api/validate-token")
async def validate_token(authorization: str = Header(None)):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Invalid token")
    
    try:
        token = authorization.split(" ")[1]
        jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        return {"valid": True}
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")
    

#print 
@app.get("/tickets/print")
async def get_tickets_for_printing(token: str = Depends(verify_token)):
    try:
        with sqlite3.connect(TICKETS_DB) as conn:
            cursor = conn.cursor()
            tickets = {
                'trib_normal': [],
                'trib_discounted': [],
                'sitz_normal': [],
                'sitz_discounted': []
            }
            
            # Get tribune tickets
            cursor.execute("""
                SELECT TicketNumber, Name_Ticketinhaber, Email_Ticketinhaber, 
                       Bestell_ID, Discounted, qrcode
                FROM trib
                ORDER BY TicketNumber
            """)
            for row in cursor.fetchall():
                category = 'trib_discounted' if row[4] else 'trib_normal'
                tickets[category].append({
                    'name': row[1],
                    'email': row[2],
                    'order_id': row[3],
                    'qrcode': row[5]  # Include QR code
                })
            
            # Get seated/standing tickets
            cursor.execute("""
                SELECT TicketNumber, Name_Ticketinhaber, Email_Ticketinhaber, 
                       Bestell_ID, Discounted, qrcode
                FROM sitz
                ORDER BY TicketNumber
            """)
            for row in cursor.fetchall():
                category = 'sitz_discounted' if row[4] else 'sitz_normal'
                tickets[category].append({
                    'name': row[1],
                    'email': row[2],
                    'order_id': row[3],
                    'qrcode': row[5]  # Include QR code
                })
                
            return tickets
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))